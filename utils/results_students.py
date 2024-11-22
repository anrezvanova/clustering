import pandas as pd
import numpy as np
from tqdm.autonotebook import tqdm
import re
from openpyxl import load_workbook


# --- БЛОК 1: Агрегация баллов ---
def get_students_from_file(file_path, excluded_students):
    """Загружает студентов из Excel-файла, исключая тестовые записи."""
    data = pd.read_excel(file_path)[['status', 'fullname']]
    students = data[data['status'] != 'teacher']['fullname']
    students = students[~students.isin(excluded_students)]
    
    return list(students)

def aggregate_scores(students, task_files, good_cols=None):
    """Выполняет агрегацию баллов студентов по загруженным файлам заданий."""
    result_table = pd.DataFrame(index=students, columns=pd.MultiIndex.from_arrays([[], []]))
    

    for task_file in tqdm(task_files):
        try:
            # Парсим баллы студентов
            table = pd.read_excel(task_file, sheet_name='Студенты')
            
            # Используем good_cols, если он передан, иначе выбираем 'Студент' и колонки начиная с D
            if good_cols:
                actual_cols = ['Студент'] + [col for col in good_cols if col in table.columns]
            else:
                actual_cols = ['Студент'] + table.columns[3:].tolist()
            table = table[actual_cols].iloc[5:-1].set_index('Студент') # оставляем нужные колонки
            table = table.loc[:, ~table.isna().all()] # отбрасываем пустые колонки, такое бывает

            # Форматируем названия колонок
            task_name = task_file.name.split(' [')[0]
            table.columns = [[task_name] * len(table.columns), table.columns]
            result_table = result_table.join(table, how='left').loc[students]
        except Exception as e:
            print(f"Не удалось обработать файл {task_file.name}: {e}")

    
    return result_table 



def aggregate_max_ball_table(task_files, good_cols=None):
    """
    Создает таблицу, где строки — это типы сумм, а столбцы — файлы.
    Заполняет только те суммы, которые присутствуют, оставляя NaN для отсутствующих значений.
    """
    # Словарь для сбора данных
    sum_dict = {}
    file_order = [task_file.name for task_file in task_files]
    # Проходим по каждому файлу
    for task_file in tqdm(task_files):
        try:
            # Считываем файл
            max_ball = pd.read_excel(
                task_file, 
                sheet_name='Список задач', 
                skiprows=3, 
                usecols="C:D", 
                header=None
            ).dropna(how='all')
            print(max_ball)
            max_ball.columns = ['Сумма', 'Значение']

            file_name = task_file.name

            # Обрабатываем каждую строку файла
            for _, row in max_ball.iterrows():
                sum_type = row['Сумма']
                value = row['Значение']
                
                # Пропускаем, если сумма не в good_cols (если указан)
                if good_cols and sum_type not in good_cols:
                    continue
                # Инициализируем вложенные словари, если их еще нет
                if sum_type not in sum_dict:
                    sum_dict[sum_type] = {}
                
                # Записываем значение для текущего файла
                sum_dict[sum_type][file_name] = value

        except Exception as e:
            print(f"Ошибка при обработке файла {task_file.name}: {e}")

    # Преобразуем данные в DataFrame
    max_ball_table = pd.DataFrame(sum_dict).T

    # Сортируем строки по good_cols (если указаны)
    if good_cols:
        max_ball_table = max_ball_table.loc[good_cols]

    # Упорядочиваем столбцы в соответствии с порядком файлов
    max_ball_table = max_ball_table.reindex(columns=file_order)

    # Удаляем столбцы с полностью NaN значениями
    max_ball_table = max_ball_table.dropna(axis=1, how='all')
    
    
    return max_ball_table



# --- БЛОК 2: Обработка результатов вопросов ---
def filter_question_files(uploaded_files, excluded_files=None):
    """
    Фильтрует загруженные файлы, исключая файлы с именами из списка excluded_files.

    Parameters:
    uploaded_files (list): Список файлов, загруженных через st.file_uploader.
    excluded_files (list): Список исключаемых файлов. По умолчанию содержит файлы для исключения.

    Returns:
    list: Отфильтрованные файлы.
    """
    if excluded_files is None:
        excluded_files = ['Оценивание занятий.xlsx', 'Посещаемость.xlsx']
    
    # Преобразуем список файлов в массив NumPy для более быстрой фильтрации
    file_names = np.array([file.name for file in uploaded_files])  # Массив имен файлов
    excluded_files = np.array(excluded_files)  # Массив исключаемых файлов
    
    # Применяем фильтрацию с помощью np.in1d
    mask = ~np.in1d(file_names, excluded_files)  # Создаем маску, исключающую указанные файлы
    file_names = list(file_names[mask])
    # Возвращаем отфильтрованные файлы
    return [uploaded_files[i] for i in range(len(uploaded_files)) if mask[i]]

def find_author(question_id, answers_list, question_id_len=4):
    answers_list = np.array(answers_list)
    question_txt_name = answers_list[[
        (x[:question_id_len] == question_id) & (x.split('.')[-1] == 'txt')
        for x in answers_list
    ]]
    
    if len(question_txt_name) == 0:
        return "Автор не найден"
    
    question_txt_name = question_txt_name[0]    
    teacher_name = ' '.join(question_txt_name[:-4].split(' ')[2:])
    return teacher_name

def add_question_info(dict_info, question_id, answers_list, date, question_text, info="", question_id_len=4):
    teacher_name = find_author(question_id, answers_list, question_id_len=question_id_len)
    
    text = (
        f"=== Вопрос {question_id}, {teacher_name} ===\n"
        + f"Дата ответа: {date}\n{question_text}\n"
        + ((info+"\n\n") if info else "\n")
    )
    
    if teacher_name not in dict_info:
        dict_info[teacher_name] = text
    else:
        dict_info[teacher_name] += text

def process_question_files(students, file_names, question_files):
    """
    Обрабатывает файлы с вопросами и ответами, собирает результаты.
    """
    students = np.array(students)
    students = [s.lstrip() for s in students if s.strip()]
    question_id_len = 4
    unsent_questions = {}
    error_questions = {}
    result_table = pd.DataFrame(index=students)
    
    print(students)
    print(file_names)
    print(question_files)
    
    # Перебор всех загруженных файлов
    for i in range(len(file_names)):
        # Читаем только xlsx
        if file_names[i].split('.')[-1] != 'xlsx':
            continue
        question_id = file_names[i][:question_id_len] 
        
        table = pd.read_excel(question_files[i])

        # Проверяем, что значение в ячейке [1, 1] является строкой и начинается с 'не'
        if table.iloc[0, 1][:2] == 'не':
            add_question_info(
                unsent_questions,
                question_id,
                answers_list=file_names,
                date=table.iloc[1, 1],  
                question_text=table.columns[1]
            )
            continue

        max_ball = float(table.iloc[2, 1])
        
        try:
            if np.isnan(max_ball) or max_ball<=0:
                # запоминаем информацию
                add_question_info(
                    error_questions,
                    question_id,
                    answers_list=file_names,
                    date=table.iloc[1, 1],
                    question_text=table.columns[1],
                    info="Некорректный максимальный балл",
                )
                continue

            table = table.iloc[7:, [1, 3]]  # Выделяем студентов и их баллы
            table.columns = ['Студент', int(file_names[i].split(' ')[0])]
            table = table.set_index('Студент')
            table = table / max_ball  # нормировка

            result_table = result_table.join(table, how='left')
            
        except Exception as e:
            # запоминаем информацию
            add_question_info(
                error_questions,
                question_id,
                answers_list=file_names,
                date=table.iloc[1, 1],
                question_text=table.columns[1],
                info=f"Что-то пошло не так: {e}",
            )
     
   
    result_table = result_table.loc[students].fillna(0)

    
    return result_table, unsent_questions, error_questions


# --- БЛОК 3: Обработка посещаемости ---
def process_attendance(file_obj,students):
    """
    Обработка файла с посещаемостью, добавляя список студентов и группируя по преподавателям.
    
    :param file_obj: файл Excel с посещаемостью.
    :param student_list: список студентов для добавления в таблицу.
    :return: DataFrame с посещаемостью.
    """
    # Загрузка файла Excel через openpyxl
    # Загрузка файла Excel через openpyxl
    # Загрузка файла Excel через openpyxl
    wb = load_workbook(file_obj, data_only=True)
    sheet = wb.active

    # Создаем новую таблицу, обрабатывая объединенные ячейки
    data = []
    for row in sheet.iter_rows():
        new_row = []
        for cell in row:
            if cell.coordinate in sheet.merged_cells:  # Проверяем, часть ли ячейка объединенного диапазона
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        new_row.append(merged_range.start_cell.value)
                        break
            else:
                new_row.append(cell.value)
        data.append(new_row)

    # Преобразуем данные в DataFrame
    df = pd.DataFrame(data)
    # Функция для очистки имени преподавателя от цифр и специальных символов
    def clean_teacher_name(name):
        # Убираем все символы, кроме букв и пробелов
        name = re.sub(r'[^\w\s]', '', name)
        # Убираем суффиксы вида _1, _2 и т.п.
        name = re.sub(r'_\d+$', '', name)
        return name.strip()
    
    
    try:
        # Шапка таблицы: преподаватели, названия занятий, даты
        first_row = df.iloc[0].fillna("Не указано").tolist()  # Преподаватели
        second_row = df.iloc[1].fillna("Не указано").tolist()  # Названия занятий
        third_row = pd.to_datetime(df.iloc[2], errors='coerce').dt.strftime('%Y-%m-%d').fillna("Не указано").tolist()  # Даты
        

        # Очистим имена преподавателей в первой строке
        cleaned_first_row = [clean_teacher_name(prof) if prof != "Не указано" else prof for prof in first_row[1:]]

        # Порядок преподавателей на основе их первого появления (уже очищенных имен)
        teacher_order = {prof: idx for idx, prof in enumerate(cleaned_first_row) if prof != "Не указано"}

        # Формируем MultiIndex из трех уровней
        multi_index = []
        seen = set()
        
        for prof, session, date in zip(first_row[1:], second_row[1:], third_row[1:]):
            cleaned_prof = clean_teacher_name(prof)
            # Проверка, является ли дата строкой, и если да, то пробуем её привести к нужному формату
            
            # Если значение не строка, проверяем на дату
            try:
                date = pd.to_datetime(date, errors='coerce')
                if pd.notna(date):
                    date_str = date.strftime('%d.%m')  # Форматируем дату в "день.месяц"
                else:
                    date_str = "Не указано"
            except Exception:
                date_str = "Не указано"
            unique_combination = (cleaned_prof, session, date_str)
            counter = 1
            while unique_combination in seen:
                unique_combination = (f"{prof}_{counter}", f"{session}_{counter}", f"{date}_{counter}")
                counter += 1
            seen.add(unique_combination)
            multi_index.append(unique_combination)
        multi_index.sort(key=lambda x: (teacher_order.get(clean_teacher_name(x[0]), float('inf')), pd.to_datetime(x[2], errors='coerce')))
        # Подготовка данных
        students_data = df.iloc[3:].reset_index(drop=True)
        
        students_data['Автор'] = students_data.iloc[:, 0].str.strip()  # Добавляем столбец с именами студентов
        students_data = students_data.drop(columns=students_data.columns[0])  # Убираем лишний первый столбец

        # Добавляем список студентов
        new_students = pd.DataFrame({'Автор': students})
        students_data = pd.merge(new_students, students_data, on='Автор', how='left')

        # Устанавливаем MultiIndex для столбцов
        columns = [('Автор', '', '')] + multi_index
        students_data.columns = pd.MultiIndex.from_tuples(columns)
         # Убедимся, что все студенты находятся в индексе
        students_data = students_data.set_index('Автор')
        
        
       
        # Фильтрация столбцов по преподавателям
        

        # Применение фильтрации
       
        # Возвращаем единую таблицу
        return students_data

    except Exception as e:
        raise ValueError(f"Ошибка при обработке файла: {e}")