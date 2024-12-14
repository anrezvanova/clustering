# Основные библиотеки
import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from openpyxl.styles import Font, PatternFill

# Импорты пользовательских модулей
from utils.clustering import Clusterer  # Кластеризация данных
from utils.ranking import Ranker  # Ранжирование данных


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

def find_task_sheets(xls):
    """Находит все листы, начинающиеся с "Задача" и заканчивающиеся цифрой."""
    task_sheets = [sheet for sheet in xls.sheet_names if re.match(r'Задача \d+', sheet)]
    return task_sheets

def clean_column_headers(df):
    """Обрабатывает заголовки таблицы, используя 5-ю строку как заголовки и удаляя первые 5 строк."""
    df.columns = df.iloc[4]
    df = df.drop([0, 1, 2, 3, 4]).reset_index(drop=True)

    return df

def get_cell_coordinates(row, col):
    """
    Преобразует индексы строки и колонки в координаты ячейки в формате Excel.
    
    :param row: Индекс строки (целое число).
    :param col: Индекс колонки (целое число).
    :return: Строка с координатами ячейки в формате Excel (например, "A1", "B2", "AA10").
    """
    col_letter = ""
    while col >= 0:
        col_letter = chr(col % 26 + ord('A')) + col_letter
        col = col // 26 - 1

    return f"{col_letter}{row + 7}"

def make_column_names_unique(df):
    """Преобразует дублирующиеся имена колонок в уникальные."""
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        cols[cols[cols == dup].index] = [f"{dup}_{i}" if i != 0 else dup for i in range(sum(cols == dup))]
    df.columns = cols
    return df
     
def find_comment_columns(df, base_columns):
    """
    Ищет все колонки, связанные с комментариями, включая уникальные названия.

    :param df: DataFrame с данными.
    :param base_columns: Базовые названия колонок для поиска (например, "Комментарий" и "Индивидуальный комментарий").
    :return: Список найденных колонок.
    """
    comment_columns = []
    for col in df.columns:
        for base_col in base_columns:
            if col.startswith(base_col):  # Поиск колонок, которые начинаются с базового имени
                comment_columns.append(col)
    return comment_columns

def find_columns_in_sheets(file_path, target_columns):
    """
    Ищет целевые колонки на листах Excel, которые соответствуют шаблону "Задача N".

    :param file_path: Путь к Excel файлу.
    :param target_columns: Целевые колонки для поиска.
    :return: Словарь с данными из найденных колонок по листам.
    """
    xls = pd.ExcelFile(file_path)
    task_sheets = find_task_sheets(xls)
    
    all_data = []
    cell_coordinates = {}

    for sheet in task_sheets:
        print(f"\nЛист: {sheet}")
        df = pd.read_excel(xls, sheet_name=sheet)
        df = clean_column_headers(df)
        
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                cell_value = df.iat[row, col]
                coord = get_cell_coordinates(row, col)  
                cell_coordinates[(row, col)] = coord  
                df.iat[row, col] = f"{cell_value} + {coord}"
       
        if all(col in df.columns for col in target_columns):
            df['Задача'] = sheet  
            df = df[target_columns + ['Задача']]
            if df.columns.duplicated().any():
                df = make_column_names_unique(df)
                stop_bot_index = df[df.apply(lambda row: row.astype(str).str.contains('STOP BOT').any(), axis=1)].index
                if len(stop_bot_index) > 0:
                    df = df.iloc[:stop_bot_index[0]]  
            all_data.append(df)     
        else:
            print(f"Не все колонки найдены на листе: {sheet}")

    if all_data:
        combined_data = pd.concat(all_data, ignore_index=True)
        if not combined_data.index.is_unique:
            print("Обнаружены дублирующиеся индексы после объединения данных.")
            combined_data = combined_data.reset_index(drop=True)
        return combined_data, cell_coordinates
    else:
        return pd.DataFrame(), {}

def creating_dictionary(df, base_columns_i, base_columns_o, cell_coordinates):
    """
    Создает словарь отфильтрованных комментариев с их исходными координатами.

    :param df: DataFrame с данными.
    :param base_columns_i: Базовые названия колонок с индивидуальными комментариями.
    :param base_columns_o: Базовые названия колонок с общими комментариями.
    :return: Словарь с отфильтрованными комментариями.
    """

    df.columns = df.columns.str.strip()
    df = make_column_names_unique(df)
    comment_columns_i = find_comment_columns(df, [base_columns_i])
    comment_columns_o = find_comment_columns(df, [base_columns_o])

    # Заполнение пропущенных значений и удаление строк с короткими или пустыми комментариями
    for col in comment_columns_i + comment_columns_o:
        df.loc[:, col] = df[col].fillna('').str.strip()
    df = df[~df[comment_columns_i + comment_columns_o].apply(lambda row: all(len(val) <= 5 for val in row), axis=1)]

    # Обрабатываем комментарии и сохраняем их исходные координаты
    comment_rows = []
    for index, row in df.iterrows():
        for col in comment_columns_o + comment_columns_i:
            if row[col]:
                comment_rows.append({
                    'Ячейка': cell_coordinates[(index, df.columns.get_loc(col))],
                    'Студент': row['Студент'],
                    'Проверяющий': row['Проверяющий'],
                    'Задача': row['Задача'],
                    'Комментарии': "*" + row[col] if col in comment_columns_o else row[col],
                    'Тип комментария': "Общий комментарий" if col in comment_columns_o else "Индивидуальный комментарий",
                })
    comments_df = pd.DataFrame(comment_rows)

    experts_comments_dict = {}
    for _, row in comments_df.iterrows():
        if '+' in row['Комментарии'] and pd.notna(row['Комментарии'].split('+')[0].strip()):
            student = row['Студент']
            reviewer = row['Проверяющий']
            comment_key = row['Комментарии'].split('+')[1].strip()
            comment_value = row['Комментарии'].split('+')[0].strip()
            
            if comment_value.lower() != 'nan' and '*nan' not in comment_value.lower():
                if student not in experts_comments_dict:
                    experts_comments_dict[student] = {'Проверяющий': reviewer, 'Комментарии': {}}
                experts_comments_dict[student]['Комментарии'][comment_key] = comment_value
    
    return experts_comments_dict

def clustering(experts_comments_dict: dict[str, any]) -> pd.DataFrame:
    '''Кластеризует комментарии экспертов и возвращает информацию о кластерах'''
    clusterer = Clusterer()  # Создаем объект без параметров
    comments, keys, students, reviewers = [], [], [], []

    for student, data in experts_comments_dict.items():
        reviewer = data['Проверяющий']
        for cell_key, com in data['Комментарии'].items():
            if isinstance(com, str) and not pd.isna(com):
                comments.append(com.strip())
                keys.append(cell_key)
                students.append(student)
                reviewers.append(reviewer)
    
    # Кластеризация
    labels = clusterer.cluster(comments, eps=15, min_samples=2)

    if len(labels) == 0:
        print("Кластеризация не выявила никаких кластеров.")
        return pd.DataFrame({
            'Ячейка': [], 'Студент': [], 'Проверяющий': [], 'Комментарии': [], 'Кластер': []
        })

    # Отбираем крупные кластеры
    unique = np.unique(labels, return_counts=True)
    big_cluster_ids = [i for (i, count) in zip(*unique) if count > 2]  
    labels = [l if l in big_cluster_ids else -1 for l in labels]

    if not any(label != -1 for label in labels):
        print("Все кластеры являются шумом. Нет подходящих кластеров.")
        return pd.DataFrame({
            'Ячейка': [], 'Студент': [], 'Проверяющий': [], 'Комментарии': [], 'Кластер': []
        })
    
    min_len = min(len(comments), len(labels), len(keys), len(students), len(reviewers))
    filtered_comments_labels_keys = [
        (comments[i], labels[i], keys[i], students[i], reviewers[i])
        for i in range(min_len) if labels[i] != -1
    ]

    if not filtered_comments_labels_keys:
        return pd.DataFrame({
            'Ячейка': [], 'Студент': [], 'Проверяющий': [], 'Комментарии': [], 'Кластер': []
        })

    # Разворачиваем фильтрованные данные
    filtered_comments, filtered_labels, filtered_keys, filtered_students, filtered_reviewers = zip(*filtered_comments_labels_keys)

    comments_df = pd.DataFrame({
        'Ячейка': filtered_keys,
        'Студент': filtered_students,
        'Проверяющий': filtered_reviewers,
        'Комментарии': filtered_comments,
        'Кластер': filtered_labels
    })

    # Присваиваем уникальные номера кластерам
    unique_labels = {label: i + 1 for i, label in enumerate(sorted(set(filtered_labels))) if label != -1}
    comments_df['Кластер'] = comments_df['Кластер'].map(unique_labels)

    if comments_df.empty:
        print("Кластеров не было получено - не были найдены схожие комментарии.")
        return pd.DataFrame({
            'Ячейка': [], 'Студент': [], 'Проверяющий': [], 'Комментарии': [], 'Кластер': []
        })
    
    # Ранжирование
    ranker = Ranker()

    try:
        # Сортируем комментарии внутри каждого кластера
        clustered_df = comments_df.groupby('Кластер', group_keys=False).apply(
            lambda group: group.iloc[ranker.rank(group['Комментарии'].tolist())]
        ).reset_index(drop=True)
    except TypeError as e:
        print("Ошибка при ранжировании: возможно, проблема с кластеризацией.", e)
        return comments_df  

    return clustered_df.sort_values(by='Кластер').reset_index(drop=True)


def create_formatted_dataframe(dataframes):
    '''Создает финальную таблицу из списка датафреймов с кластеризацией'''

    if not dataframes:
        print("Нет данных для формирования итоговой таблицы.")
        return pd.DataFrame(columns=["Номер", "Студент", "Проверяющий", "Комментарии"])
    
    formatted_dfs = []
    task_numbers = [df["Задача"].tolist()[0] for df in dataframes]

    for task, df in zip(task_numbers, dataframes):
        if df.empty:
            print(f"Пропуск пустого датафрейма для задачи {task}")
            continue
    
        # Добавляем пустую строку перед задачей (разделение)
        header_df = pd.DataFrame({ 
            "Номер": [""], 
            "Студент": [""],
            "Проверяющий": [""],
            "Комментарии": [""]
        })
        formatted_dfs.append(header_df)

        # Добавляем сам заголовок задачи
        task_header_df = pd.DataFrame({ 
            "Номер": [f"--- {task} ---"], 
            "Студент": [""],
            "Проверяющий": [""],
            "Комментарии": [""]
        })
        formatted_dfs.append(task_header_df)
        
        # Обработка заголовков кластеров
        for cluster_num, cluster_df in df.groupby("Кластер"):
            cluster_header = pd.DataFrame({
                "Номер": [f"Кластер {cluster_num}:"],  # Заголовок для каждого кластера
                "Студент": [""],
                "Проверяющий": [""],
                "Комментарии": [""]
            })
            formatted_dfs.append(cluster_header)
            formatted_dfs.append(cluster_df.reset_index(drop=True))

    if not formatted_dfs:
        print("Нет данных после обработки кластеров.")
        return pd.DataFrame(columns=["Номер", "Студент", "Проверяющий", "Комментарии"])

    final_df = pd.concat(formatted_dfs, ignore_index=True)
    final_df = final_df.drop(columns=["Кластер","Задача"])
    final_df.fillna("", inplace=True)
    if "Студент" in final_df.columns:
        final_df["Студент"] = final_df["Студент"].apply(lambda x: x.split("+")[0] if isinstance(x, str) else x)
    if "Проверяющий" in final_df.columns:
        final_df["Проверяющий"] = final_df["Проверяющий"].apply(lambda x: x.split("+")[0] if isinstance(x, str) else x)
    return final_df


    
def save_results_to_excel(df, base_file_name):
    """
    Создаёт Excel с результатами и возвращает буфер для скачивания.
    
    :param df: DataFrame с результатами.
    :param base_file_name: Базовое имя файла (без расширения).
    :return: Буфер с Excel-файлом.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.fillna("", inplace=True)
        df.to_excel(writer, index=False, sheet_name='Результаты')
        worksheet = writer.sheets['Результаты']

        for col in worksheet.columns:
            max_length = max((len(str(cell.value)) for cell in col), default=0)
            column_letter = col[0].column_letter
            worksheet.column_dimensions[column_letter].width = max_length + 2

        for row in worksheet.iter_rows():
            for cell in row:
                if 'Кластер' in str(cell.value):
                    cell.font = Font(bold=True)
                if '---' in str(cell.value):
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    cell.font = Font(bold=True)

    buffer.seek(0)
    return buffer

def main(file_path, student_column, reviewer_column, comment_column_i, comment_column_o):
    
    target_columns = [student_column, reviewer_column, comment_column_i, comment_column_o]
    all_task_data, cell_coordinates = find_columns_in_sheets(file_path, target_columns)
    task_names = all_task_data['Задача'].unique()

    list_clustered_info = []
    for task in task_names:
        task_data = all_task_data[all_task_data['Задача'] == task]
        if not task_data.empty:
            print(f"Кластеризация для задачи {task}")
            experts_comments_dict = creating_dictionary(task_data, comment_column_i, comment_column_o, cell_coordinates)
           
            if not experts_comments_dict:  
                print("Внимание: experts_comments_dict пустой. Пропускаем кластеризацию.")
                continue
            clustered_info = clustering(experts_comments_dict)
           
            if not clustered_info.empty:
                clustered_info['Задача'] = task
                list_clustered_info.append(clustered_info)
            else:
                print("Данные не были найдены.")
    final_df = create_formatted_dataframe(list_clustered_info)

    
    return final_df







